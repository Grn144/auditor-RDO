"""Testes do preenchimento da planilha de medição: casamento de linhas
(grupo+item) com as tarefas do app e cálculo do incremento por rodada.

As colunas QT./MEDIÇÃO 0N e a linha de cabeçalho NÃO são fixas — dois
arquivos reais de clientes diferentes (AXA e Magalu) vieram com layouts
diferentes (linha do cabeçalho e nº de colunas descritivas antes de
QT. mudam), então os testes cobrem explicitamente mais de um layout."""
import openpyxl
import pytest

import planilha_medicao as pm

# Layout da planilha real da AXA: cabeçalho na linha 7, QT. da rodada 1
# na coluna P (16) — usado como padrão nos testes que não são sobre
# variação de layout.
_HEADER_ROW_PADRAO = 7
_RODADA_COLS_PADRAO = {1: 16, 2: 19, 3: 22, 4: 25}


def _wb_com_linhas(linhas, header_row=_HEADER_ROW_PADRAO, rodada_cols=_RODADA_COLS_PADRAO):
    """Monta um workbook mínimo no formato da planilha de medição: uma
    linha de cabeçalho (coluna A="ITEM", e para cada rodada um par de
    células "QT."/"MEDIÇÃO 0N"), seguida das linhas de item — coluna
    A=grupo, B=item, C=descrição, e valores extras (ex.: QT. de rodadas
    já lançadas) nas colunas indicadas.

    `linhas`: lista de dicts {row, a, b, c, valores: {coluna: valor}}."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=header_row, column=1, value="ITEM")
    ws.cell(row=header_row, column=2, value="N°")
    ws.cell(row=header_row, column=3, value="DESCRIÇÃO")
    for numero, col_qt in rodada_cols.items():
        ws.cell(row=header_row, column=col_qt, value="QT.")
        ws.cell(row=header_row, column=col_qt + 1, value=f"MEDIÇÃO 0{numero}")
    for linha in linhas:
        ws.cell(row=linha["row"], column=1, value=linha["a"])
        ws.cell(row=linha["row"], column=2, value=linha["b"])
        ws.cell(row=linha["row"], column=3, value=linha.get("c", ""))
        for coluna, valor in linha.get("valores", {}).items():
            ws.cell(row=linha["row"], column=coluna, value=valor)
    return wb


def test_preenche_incremento_quando_nao_ha_rodada_anterior():
    wb = _wb_com_linhas([{"row": 8, "a": "D", "b": "1", "c": "ITEM D1"}])
    atividades = [{"grupo": "D", "codigo": "1", "porcentagem": 90, "quantidade": 1}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=1)

    assert wb.active.cell(row=8, column=16).value == 0.9  # coluna P = QT. medição 01
    assert avisos == []


def test_calcula_so_o_incremento_desde_a_rodada_anterior():
    """Exemplo real do usuário: medição 1 lançou 0,8 de um item que
    agora está 100% — a medição 2 recebe só a diferença (0,2)."""
    wb = _wb_com_linhas([{"row": 8, "a": "A", "b": "1", "c": "ITEM A1",
                           "valores": {16: 0.8}}])
    atividades = [{"grupo": "A", "codigo": "1", "porcentagem": 100, "quantidade": 1}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=2)

    assert wb.active.cell(row=8, column=19).value == 0.2  # coluna S = QT. medição 02
    assert avisos == []


def test_nao_escreve_e_avisa_quando_incremento_seria_negativo():
    wb = _wb_com_linhas([{"row": 8, "a": "A", "b": "1", "c": "ITEM A1",
                           "valores": {16: 0.9}}])
    atividades = [{"grupo": "A", "codigo": "1", "porcentagem": 50, "quantidade": 1}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=2)

    assert wb.active.cell(row=8, column=19).value is None
    assert len(avisos) == 1
    assert "A1" in avisos[0]


def test_avisa_item_da_planilha_sem_tarefa_correspondente_no_app():
    wb = _wb_com_linhas([{"row": 8, "a": "A", "b": "1", "c": "ITEM A1"}])

    avisos = pm.preencher_medicao(wb, atividades=[], rodada=1)

    assert len(avisos) == 1
    assert "A1" in avisos[0]


def test_avisa_tarefa_do_app_sem_linha_correspondente_na_planilha():
    wb = _wb_com_linhas([])
    atividades = [{"grupo": "B", "codigo": "2", "porcentagem": 50, "quantidade": 2}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=1)

    assert len(avisos) == 1
    assert "B2" in avisos[0]


def test_ignora_linha_de_cabecalho_de_categoria():
    """Linhas onde grupo == item (ex.: A/A) são o subtotal da categoria,
    não um item — não devem ser tratadas como item nem gerar aviso."""
    wb = _wb_com_linhas([
        {"row": 8, "a": "A", "b": "A", "c": "PRÉ-OBRA"},
        {"row": 9, "a": "A", "b": "1", "c": "ITEM A1"},
    ])
    atividades = [{"grupo": "A", "codigo": "1", "porcentagem": 100, "quantidade": 2}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=1)

    assert wb.active.cell(row=9, column=16).value == 2.0
    assert avisos == []


def test_avisa_quando_tarefa_sem_quantidade_cadastrada():
    wb = _wb_com_linhas([{"row": 8, "a": "A", "b": "1", "c": "ITEM A1"}])
    atividades = [{"grupo": "A", "codigo": "1", "porcentagem": 40, "quantidade": None}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=1)

    assert wb.active.cell(row=8, column=16).value is None
    assert len(avisos) == 1


def test_rodada_invalida_levanta_erro():
    wb = _wb_com_linhas([])
    with pytest.raises(ValueError):
        pm.preencher_medicao(wb, [], rodada=5)


def test_funciona_com_planilha_de_layout_diferente():
    """Bug real: a planilha de outro cliente (Magalu) tem o cabeçalho
    numa linha diferente e 2 colunas a menos antes do bloco de QT./
    MEDIÇÃO (menos colunas descritivas que a da AXA) — a coluna da
    rodada 1 cai em N (14), não em P (16). O preenchimento tem que
    achar a coluna certa pelo texto do cabeçalho, não por posição
    fixa, senão escreve no lugar errado (como aconteceu de verdade)."""
    header_row = 8
    rodada_cols = {1: 14, 2: 17, 3: 20, 4: 23}  # N, Q, T, W
    wb = _wb_com_linhas(
        [{"row": 9, "a": "A", "b": "1", "c": "ITEM A1"}],
        header_row=header_row, rodada_cols=rodada_cols,
    )
    atividades = [{"grupo": "A", "codigo": "1", "porcentagem": 90, "quantidade": 1}]

    avisos = pm.preencher_medicao(wb, atividades, rodada=1)

    ws = wb.active
    assert ws.cell(row=9, column=14).value == 0.9   # coluna certa (N)
    assert ws.cell(row=9, column=16).value is None  # não pode cair na coluna da AXA (P)
    assert avisos == []


def test_planilha_sem_cabecalho_reconhecivel_levanta_erro():
    wb = openpyxl.Workbook()  # workbook em branco, sem "ITEM" em lugar nenhum
    with pytest.raises(ValueError):
        pm.preencher_medicao(wb, [], rodada=1)
